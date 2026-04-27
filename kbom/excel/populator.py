"""Write extracted BOM rows into the customer's Excel template.

Targets three sheets, ALL of which are inputs to the formula chain:
- 장등록(규격,수량,옵션) — 목대 (custom frame parts), rows 4–18, cols B–F
- 상품등록 — 상품 (off-the-shelf products), rows 8+, cols A and C
- 기초입력(사양등록) — project-level assumptions

The write-target row layout matches the convention proven in the Phase 0
round-trip test: row 4 starts the 목대 list, panels first then base run.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from kbom.excel.recalc import find_soffice
from kbom.models import CabinetRow, Category


# Sheet + cell layout (proven 0-variance in Phase 0 round-trip)
JANG_SHEET = "장등록(규격,수량,옵션)"
SANG_SHEET = "상품등록"
INPUT_SHEET = "기초입력(사양등록)"

JANG_FIRST_ROW = 4   # row 4 in openpyxl 1-indexed; row 3 is header
SANG_FIRST_ROW = 8   # row 8 starts product list
INPUT_TYPE_CELL = "C4"
INPUT_UNITS_CELL = "C5"


def populate_template(
    template_path: str | Path,
    output_path: str | Path,
    rows: list[CabinetRow],
    type_label: str = "26A",
    units_count: int = 1,
) -> Path:
    """Write rows into a copy of the template, returning the populated path.

    Auto-converts legacy `.xls` to `.xlsx` if needed (openpyxl is xlsx-only).
    Caller is responsible for triggering LibreOffice recalc on the result.
    """
    src = Path(template_path)
    dst = Path(output_path)

    # If template is legacy .xls, convert to .xlsx first
    if src.suffix.lower() == ".xls":
        src = _convert_xls_to_xlsx(src, dst.parent)

    # Ensure output is .xlsx
    if dst.suffix.lower() != ".xlsx":
        dst = dst.with_suffix(".xlsx")

    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst, data_only=False)

    # Split rows by category
    mokdae = [r for r in rows if r.category == Category.MOKDAE]
    sangpum = [r for r in rows if r.category == Category.SANGPUM]

    _write_mokdae(wb, mokdae)
    _write_sangpum(wb, sangpum)
    _write_assumptions(wb, type_label=type_label, units_count=units_count)

    wb.save(dst)
    return dst


def _write_mokdae(wb: openpyxl.Workbook, rows: list[CabinetRow]) -> None:
    """Write 목대 rows into 장등록 sheet, cols B (code), C (name), D-F (W/D/H)."""
    ws = wb[JANG_SHEET]

    # Clear existing rows in the write-target range first (cols B-F only;
    # cols A and G+ contain formulas/computed values that we leave alone)
    end_row = JANG_FIRST_ROW + max(len(rows), 15)
    for r in range(JANG_FIRST_ROW, end_row + 1):
        for col in range(2, 7):  # B..F
            ws.cell(r, col).value = None

    # Write rows in given order
    for i, row in enumerate(rows):
        excel_row = JANG_FIRST_ROW + i
        ws.cell(excel_row, 2).value = row.code              # B
        ws.cell(excel_row, 3).value = row.name              # C
        ws.cell(excel_row, 4).value = row.width_mm          # D
        ws.cell(excel_row, 5).value = row.depth_mm          # E
        ws.cell(excel_row, 6).value = row.height_mm         # F


def _write_sangpum(wb: openpyxl.Workbook, rows: list[CabinetRow]) -> None:
    """Write 상품 rows into 상품등록 sheet, col A (key) and col C (name)."""
    ws = wb[SANG_SHEET]

    end_row = SANG_FIRST_ROW + max(len(rows), 6)
    for r in range(SANG_FIRST_ROW, end_row + 1):
        ws.cell(r, 1).value = None  # A — key
        ws.cell(r, 3).value = None  # C — name

    for i, row in enumerate(rows):
        excel_row = SANG_FIRST_ROW + i
        ws.cell(excel_row, 1).value = f"상품{row.name}"  # A — concatenated key
        ws.cell(excel_row, 3).value = row.name           # C — display name


def _write_assumptions(
    wb: openpyxl.Workbook, type_label: str, units_count: int
) -> None:
    """Write project-level assumptions into 기초입력 sheet."""
    ws = wb[INPUT_SHEET]
    ws[INPUT_TYPE_CELL] = type_label
    ws[INPUT_UNITS_CELL] = units_count


def _convert_xls_to_xlsx(xls_path: Path, output_dir: Path) -> Path:
    """Convert a legacy .xls file to .xlsx via headless LibreOffice.

    Returns the path to the converted .xlsx. One-time per template upload.
    """
    import subprocess
    import os

    output_dir.mkdir(parents=True, exist_ok=True)
    soffice = find_soffice()
    env = os.environ.copy()
    env.setdefault("HOME", str(Path.home()))

    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", str(output_dir),
            str(xls_path),
        ],
        capture_output=True,
        text=True,
        timeout=120,
        env=env,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice xls→xlsx conversion failed: {result.stderr or result.stdout}"
        )

    out_path = output_dir / (xls_path.stem + ".xlsx")
    if not out_path.exists():
        raise RuntimeError(f"Expected converted file at {out_path}, but not found.")
    return out_path
