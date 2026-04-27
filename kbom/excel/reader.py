"""Read computed cell values back from a recalculated Excel file.

This is the bridge from the customer's Excel engine to our web UI. We read
the populated input sheets AND the computed output sheets (원가산출표, 자재구성표,
일위대가표) and serve everything as structured Python objects to the front end.

For the cell inspector, we also retain the FORMULA strings (loaded with
data_only=False) so we can show them to the auditor on demand.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# Output sheets we care about for the BOM / cost rollup
SHEETS_OF_INTEREST = [
    "원가산출표",        # Computed cabinet list view (cols B-G)
    "자재구성표(BODY)",  # Body material breakdown
    "자재구성표(DOOR)",  # Door material breakdown
    "일위대가표",        # Unit cost breakdown
    "장등록(규격,수량,옵션)",  # Inputs (re-read for cell inspector)
    "상품등록",
]


@dataclass
class CellSnapshot:
    """One cell's value + formula, for the cell inspector."""

    sheet: str
    coord: str
    computed_value: Any | None = None
    formula: str | None = None


@dataclass
class WorkbookSnapshot:
    """All readable data from a recalculated workbook."""

    by_sheet: dict[str, dict[str, Any]] = field(default_factory=dict)
    formulas: dict[str, dict[str, str]] = field(default_factory=dict)

    def get_value(self, sheet: str, coord: str) -> Any | None:
        return self.by_sheet.get(sheet, {}).get(coord)

    def get_formula(self, sheet: str, coord: str) -> str | None:
        return self.formulas.get(sheet, {}).get(coord)

    def cell(self, sheet: str, coord: str) -> CellSnapshot:
        return CellSnapshot(
            sheet=sheet,
            coord=coord,
            computed_value=self.get_value(sheet, coord),
            formula=self.get_formula(sheet, coord),
        )


def read_workbook(xlsx_path: str | Path, sheets: list[str] | None = None) -> WorkbookSnapshot:
    """Load a recalculated xlsx and return both computed values and formulas.

    Computed values come from `data_only=True`; formulas from `data_only=False`.
    Only sheets in `sheets` are loaded (default: SHEETS_OF_INTEREST).
    """
    src = Path(xlsx_path)
    sheet_names = sheets or SHEETS_OF_INTEREST

    snap = WorkbookSnapshot()

    # Read computed values
    wb_vals = openpyxl.load_workbook(src, data_only=True)
    for sn in sheet_names:
        if sn not in wb_vals.sheetnames:
            continue
        ws = wb_vals[sn]
        cells: dict[str, Any] = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value
        snap.by_sheet[sn] = cells
    wb_vals.close()

    # Read formulas
    wb_fml = openpyxl.load_workbook(src, data_only=False)
    for sn in sheet_names:
        if sn not in wb_fml.sheetnames:
            continue
        ws = wb_fml[sn]
        formulas: dict[str, str] = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formulas[cell.coordinate] = v
        snap.formulas[sn] = formulas
    wb_fml.close()

    return snap
