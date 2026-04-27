"""Diff the recomputed (post-write + recalc) xlsx against the baseline snapshot."""
import openpyxl, json

BASELINE = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/1. LH 화성태안3 A2BL-26A.xlsx"
import sys
RECOMPUTED = sys.argv[1] if len(sys.argv) > 1 else "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/recalc/working_v2.xlsx"

wb_base = openpyxl.load_workbook(BASELINE, data_only=True)
wb_recomp = openpyxl.load_workbook(RECOMPUTED, data_only=True)

def vals(wb):
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cells = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value
        out[sn] = cells
    return out

base = vals(wb_base)
recomp = vals(wb_recomp)

print("Comparing per-sheet cell values (populated cells only)...")
print("=" * 90)

total_cells_base = sum(len(v) for v in base.values())
total_cells_recomp = sum(len(v) for v in recomp.values())
total_diff = 0
total_num_diff = 0
significant_num_diff = 0  # >1% relative diff on numeric cells

# Numeric tolerance for float comparison (formula recalc can cause tiny rounding diffs)
NUM_TOL_ABS = 0.01
NUM_TOL_REL = 0.001  # 0.1%

for sn in base:
    if sn not in recomp:
        print(f"  {sn}: sheet missing in recomputed!")
        continue
    diffs = []
    only_base = set(base[sn]) - set(recomp[sn])
    only_recomp = set(recomp[sn]) - set(base[sn])
    common = set(base[sn]) & set(recomp[sn])
    num_cell_diffs = 0
    exact_mismatches = 0
    for coord in common:
        b = base[sn][coord]
        r = recomp[sn][coord]
        if b == r:
            continue
        # Try numeric comparison
        try:
            bf, rf = float(b), float(r)
            if abs(bf - rf) <= NUM_TOL_ABS:
                continue
            if bf != 0 and abs(bf - rf) / abs(bf) <= NUM_TOL_REL:
                continue
            diffs.append((coord, b, r))
            num_cell_diffs += 1
        except (TypeError, ValueError):
            if str(b).strip() == str(r).strip():
                continue
            diffs.append((coord, b, r))
            exact_mismatches += 1
    total = len(diffs) + len(only_base) + len(only_recomp)
    total_diff += total
    total_num_diff += num_cell_diffs
    marker = "✅" if total == 0 else "⚠️ " if total < 20 else "❌"
    print(f"  {marker} {sn}:  {total:5d} diffs  (numeric={num_cell_diffs}, other={exact_mismatches}, only-base={len(only_base)}, only-recomp={len(only_recomp)})")

    # Sample a few diffs for interesting sheets
    if total > 0 and sn in ("원가산출표", "기초입력(사양등록)", "상품등록", "장등록(규격,수량,옵션)"):
        for coord, b, r in diffs[:10]:
            bs = str(b)[:40]
            rs = str(r)[:40]
            print(f"      {coord}: baseline={bs!r:<45}  recomputed={rs!r}")

print("=" * 90)
print(f"Baseline total populated cells:   {total_cells_base:>8}")
print(f"Recomputed total populated cells: {total_cells_recomp:>8}")
print(f"Total cells differing (any kind): {total_diff:>8}")
print(f"  → as % of baseline: {100*total_diff/total_cells_base:.2f}%")
