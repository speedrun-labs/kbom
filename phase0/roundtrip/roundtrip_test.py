"""Round-trip test: verify that writing our extracted rows into 원가산출표 B-G
preserves the computed values across all downstream sheets.

Procedure:
  1. Load the converted .xlsx as-is, snapshot computed values from every sheet.
  2. Make a working copy, clear 원가산출표 rows 8..28 (0-indexed 7..27) cols B-G.
  3. Write our augmented 21-row BOM into the same rows.
  4. (Caller will re-save with LibreOffice afterwards to trigger recalc.)
  5. After recalc, re-read computed values and diff vs snapshot.

This script handles the snapshot + write phases. A second script handles diff.
"""
import openpyxl, json, shutil, pathlib

BASELINE = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/1. LH 화성태안3 A2BL-26A.xlsx"
WORKING  = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/working.xlsx"
AUGMENTED = "/Users/d/Desktop/NEFS Kitchen/phase0/test3_prediction_augmented.json"
SNAPSHOT = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/snapshot_baseline.json"

# --- Phase A: snapshot baseline computed values ---
print("=== Phase A: snapshotting baseline ===")
wb_read = openpyxl.load_workbook(BASELINE, data_only=True)
baseline = {}
for sn in wb_read.sheetnames:
    ws = wb_read[sn]
    sheet_data = {}
    # Walk only populated cells to keep snapshot compact
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None:
                # Use A1 coordinate as key
                sheet_data[cell.coordinate] = cell.value
    baseline[sn] = sheet_data
    print(f"  {sn}: {len(sheet_data)} populated cells")
wb_read.close()

with open(SNAPSHOT, "w", encoding="utf-8") as f:
    json.dump(baseline, f, ensure_ascii=False, default=str)
print(f"  Saved snapshot: {SNAPSHOT}\n")

# --- Phase B: make working copy and write our rows ---
print("=== Phase B: writing extracted rows into 원가산출표 B-G ===")
shutil.copy(BASELINE, WORKING)

wb = openpyxl.load_workbook(WORKING, data_only=False)  # preserve formulas
ws = wb["원가산출표"]

# Inspect current 원가산출표 B-G for 26A rows (row index 8-28 in 1-indexed openpyxl,
# which is xlrd row 7-27 0-indexed)
print("  Current 원가산출표 B-G rows 8..28 (sampling):")
for r in range(8, 29):
    b, c, d, e, f, g = (ws.cell(r, i).value for i in (2,3,4,5,6,7))
    if any(v not in (None, "") for v in (b, c, d, e, f, g)):
        print(f"    row {r}: B={b!r:15} C={c!r:20} D={d} E={e} F={f} G={g}")

# Load our augmented prediction
with open(AUGMENTED, encoding="utf-8") as f:
    pred = json.load(f)["predicted_rows"]

# Clear rows 8..28 cols B-G
print(f"\n  Clearing rows 8..28, cols B-G...")
for r in range(8, 29):
    for col in range(2, 8):  # B..G
        ws.cell(r, col).value = None

# Write our augmented rows starting at row 8
# Schema: category(B), name(C), w(D), d(E), h(F), type(G)
# Note: original Excel uses row 8 as "구분 | 제품명 | 가로 | 세로 | 높이 | TYPE" header,
# and row 8-34 pattern. Let's match the original layout precisely:
# - rows 8..22 contain 목대 rows (15 items)
# - rows 25..30 contain 상품 rows (6 items, with subtotal between)
# For simplicity in the test, we'll write all 21 rows starting at row 8
# and leave gaps to match original layout.

# The original 원가산출표 had 목대 rows 7-21 and 상품 rows 22-27 by xlrd's 0-indexing
# which is openpyxl rows 8-22 and 23-28. Let's use that same layout.
def write_row(r, item):
    ws.cell(r, 2).value = item["category"]
    ws.cell(r, 3).value = item["name"]
    ws.cell(r, 4).value = item.get("w")
    ws.cell(r, 5).value = item.get("d")
    ws.cell(r, 6).value = item.get("h")
    ws.cell(r, 7).value = item.get("type", "26A")

mokdae = [p for p in pred if p["category"] == "목대"]
sangpum = [p for p in pred if p["category"] == "상품"]

row = 8
for item in mokdae:
    write_row(row, item); row += 1
# Skip if fewer than 15 rows
for item in sangpum:
    write_row(row, item); row += 1

print(f"  Wrote {len(mokdae)} 목대 + {len(sangpum)} 상품 rows starting at row 8")

wb.save(WORKING)
print(f"  Saved working copy: {WORKING}")
print("\n  NEXT: re-save with LibreOffice to trigger formula recalc")
