"""Round-trip v2 — write to the CORRECT input sheets (장등록 + 상품등록 + 기초입력),
not the computed view (원가산출표).

Procedure:
  1. Start from a fresh copy of the baseline xlsx (no prior corruption).
  2. Clear 장등록 rows 4..18 cols B-F (15 목대 rows).
  3. Clear 상품등록 rows 8..13 col C (6 상품 rows).
  4. Write our augmented 21-row BOM into the correct cells.
  5. Save; let LibreOffice recalc when we reopen.
"""
import openpyxl, json, shutil

BASELINE = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/1. LH 화성태안3 A2BL-26A.xlsx"
WORKING  = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/working_v2.xlsx"
AUGMENTED = "/Users/d/Desktop/NEFS Kitchen/phase0/test3_prediction_augmented.json"

# Name → 장등록 short code. Derived from the baseline's existing entries.
NAME_TO_CODE = {
    "벽장판넬":      "WP",
    "일반벽장":      "W",
    "밑장휠라":      "BI",
    "찬넬밑장":      "C",
    "찬넬렌지밑장":   "CR",
    "찬넬3단서랍밑장": "CD3",
    "찬넬씽크밑장":   "CS",
    "밑장판넬":      "BP",
    "장식판":        "FA",
    "걸레받이":      "PL",
}

shutil.copy(BASELINE, WORKING)
wb = openpyxl.load_workbook(WORKING, data_only=False)

# --- 장등록 — 목대 rows ---
ws_jang = wb["장등록(규격,수량,옵션)"]
# Clear rows 4..18 cols B-F (keep col A = index, and G-M = computed formulas)
for r in range(4, 19):
    for c in range(2, 7):  # B=2, C=3, D=4, E=5, F=6
        ws_jang.cell(r, c).value = None

# Load augmented prediction
with open(AUGMENTED, encoding="utf-8") as f:
    pred = json.load(f)["predicted_rows"]

mokdae = [p for p in pred if p["category"] == "목대"]
sangpum = [p for p in pred if p["category"] == "상품"]

print(f"Writing {len(mokdae)} 목대 rows into 장등록 (rows 4..{3+len(mokdae)})")
for i, item in enumerate(mokdae):
    r = 4 + i
    code = NAME_TO_CODE.get(item["name"], "?")
    ws_jang.cell(r, 2).value = code
    ws_jang.cell(r, 3).value = item["name"]
    ws_jang.cell(r, 4).value = int(item["w"]) if item.get("w") else None
    ws_jang.cell(r, 5).value = int(item["d"]) if item.get("d") else None
    ws_jang.cell(r, 6).value = int(item["h"]) if item.get("h") else None
    print(f"  r{r}: {code:4} {item['name']:<15} W={item.get('w'):<5} D={item.get('d'):<4} H={item.get('h'):<5}")

# --- 상품등록 — 상품 rows ---
ws_sang = wb["상품등록"]
# Baseline has rows 8-13 populated with product names in col C
# We'll clear col C for those rows and rewrite
for r in range(8, 14):
    ws_sang.cell(r, 3).value = None
    # Col A = product key (e.g. "상품행거레일") — keep/regenerate

print(f"\nWriting {len(sangpum)} 상품 rows into 상품등록 (rows 8..{7+len(sangpum)})")
for i, item in enumerate(sangpum):
    r = 8 + i
    name = item["name"]
    ws_sang.cell(r, 3).value = name
    # Column A is the concatenated key — the formula in 원가산출표 uses OFFSET so
    # we keep col A as is (it was computed); but we should verify.
    print(f"  r{r}: {name}")

wb.save(WORKING)
print(f"\nSaved: {WORKING}")
print("\nNext: run `soffice --headless --calc --convert-to xlsx` to force recalc, then diff.")
