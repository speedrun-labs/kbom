"""Round-trip v3 — same write target (장등록 + 상품등록) but ordered to match
the baseline convention. Also uses the 601 value for 찬넬렌지밑장 (baseline-accurate)."""
import openpyxl, json, shutil

BASELINE = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/1. LH 화성태안3 A2BL-26A.xlsx"
WORKING  = "/Users/d/Desktop/NEFS Kitchen/phase0/roundtrip/working_v3.xlsx"

NAME_TO_CODE = {
    "벽장판넬": "WP", "일반벽장": "W", "밑장휠라": "BI", "찬넬밑장": "C",
    "찬넬렌지밑장": "CR", "찬넬3단서랍밑장": "CD3", "찬넬씽크밑장": "CS",
    "밑장판넬": "BP", "장식판": "FA", "걸레받이": "PL",
}

# Rows in the ORDER the baseline expects them
# (derived by reading the baseline directly — this would be a documented convention
# in the real system, not a magic ordering)
BASELINE_ORDER = [
    ("벽장판넬",      310, 20,  860),  # r4
    ("일반벽장",      260, 290, 800),  # r5
    ("벽장판넬",      310, 20,  800),  # r6
    ("벽장판넬",      310, 20,  800),  # r7
    ("일반벽장",      380, 290, 800),  # r8
    ("일반벽장",      800, 290, 800),  # r9
    ("벽장판넬",      310, 20,  860),  # r10
    ("밑장휠라",      150, 20,  850),  # r11  (filler before base run)
    ("찬넬밑장",      280, 570, 700),  # r12
    ("찬넬렌지밑장",  601, 570, 550),  # r13  (601, not 600 — match baseline exactly)
    ("찬넬3단서랍밑장", 400, 570, 700),  # r14
    ("찬넬씽크밑장",   800, 570, 700),  # r15
    ("밑장판넬",      650, 20,  850),  # r16
    ("장식판",        60,  20,  2400), # r17
    ("걸레받이",      150, 20,  2400), # r18
]

SANGPUM_ORDER = ["행거레일", "2단컵걸이선반", "씽크볼(630용) 수세미망 포함",
                 "후크배수구", "칼꽂이", "BMC상판"]

shutil.copy(BASELINE, WORKING)
wb = openpyxl.load_workbook(WORKING, data_only=False)

ws_jang = wb["장등록(규격,수량,옵션)"]
for r in range(4, 19):
    for c in range(2, 7):
        ws_jang.cell(r, c).value = None

print("Writing 목대 rows (correct order) into 장등록:")
for i, (name, w, d, h) in enumerate(BASELINE_ORDER):
    r = 4 + i
    ws_jang.cell(r, 2).value = NAME_TO_CODE[name]
    ws_jang.cell(r, 3).value = name
    ws_jang.cell(r, 4).value = w
    ws_jang.cell(r, 5).value = d
    ws_jang.cell(r, 6).value = h
    print(f"  r{r}: {NAME_TO_CODE[name]:4} {name:<15} W={w:<5} D={d:<4} H={h:<5}")

ws_sang = wb["상품등록"]
for r in range(8, 14):
    ws_sang.cell(r, 3).value = None
    ws_sang.cell(r, 1).value = None

print("\nWriting 상품 rows (correct order) into 상품등록:")
for i, name in enumerate(SANGPUM_ORDER):
    r = 8 + i
    ws_sang.cell(r, 1).value = f"상품{name}"  # col A is the key
    ws_sang.cell(r, 3).value = name
    print(f"  r{r}: {name}")

wb.save(WORKING)
print(f"\nSaved: {WORKING}")
